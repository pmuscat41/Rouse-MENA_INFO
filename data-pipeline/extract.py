"""
Extract patent filing requirements from the Excel pricelists in
INTA/Price_Lists/ and emit assets/data/requirements.json.

Robustness strategy
-------------------
- Country identity is taken from the filename, not the sheet name (sheet names
  are inconsistent: "UAE - Patents", "Saudi Arabia 2026", "Egypt " with a
  trailing space, etc.).
- The formality table is located by searching for known document-name keywords
  in the workbook, not by fixed row/column positions. This handles layout
  variation across files (Israel and Turkey have a squashed header; UAE and
  Saudi have wider fee tables that push the formality table further right).
- The "Minimum requirements for filing" info panel is located by searching for
  the "Applicant" header keyword.
- Typos are normalised silently ("Appostile" -> "Apostille",
  "Certifed" -> "Certified", "requried" -> "required",
  Excel _x000D_ artefacts and stray whitespace).

Coordinate convention
---------------------
host_coordinates is stored as [longitude, latitude] to match D3's geoMercator
input convention. The earlier draft of instructions.md mixed this up; we follow
the section 11 usage, not section 5.
"""

from __future__ import annotations

import json
import re
import sys
from datetime import date, datetime, timezone
from pathlib import Path

import openpyxl


# ----------------------------------------------------------------------------
# Paths
# ----------------------------------------------------------------------------

SCRIPT_DIR = Path(__file__).resolve().parent
REPO_ROOT = SCRIPT_DIR.parent
PRICE_LISTS_DIR = REPO_ROOT / "INTA" / "Price_Lists"
OUTPUT_FILE = REPO_ROOT / "assets" / "data" / "requirements.json"
LOG_FILE = SCRIPT_DIR / "extract.log"


# ----------------------------------------------------------------------------
# Country / regional-system metadata
# ----------------------------------------------------------------------------
#
# Keyed by the *cleaned filename* (lowercase, year prefix stripped, " - Patent"
# suffix stripped, internal whitespace collapsed). See filename_to_key() below.

COUNTRIES: dict[str, dict] = {
    # Middle East
    "uae": {
        "id": "uae",
        "name": "United Arab Emirates",
        "region": "mideast",
        "iso_numeric": "784",
        "iso_alpha3": "ARE",
        "office": "UAE Ministry of Economy",
        "filing_language": "English (Arabic translation required)",
        "routes": ["Paris", "PCT National Phase"],
    },
    "saudi arabia": {
        "id": "saudi-arabia",
        "name": "Saudi Arabia",
        "region": "mideast",
        "iso_numeric": "682",
        "iso_alpha3": "SAU",
        "office": "Saudi Authority for Intellectual Property (SAIP)",
        "filing_language": "Arabic",
        "routes": ["Paris", "PCT National Phase"],
    },
    "bahrain": {
        "id": "bahrain",
        "name": "Bahrain",
        "region": "mideast",
        "iso_numeric": "048",
        "iso_alpha3": "BHR",
        "office": "Industrial Property Office, Ministry of Industry, Commerce and Tourism",
        "filing_language": "Arabic (English accepted with translation)",
        "routes": ["Paris", "PCT National Phase"],
    },
    "kuwait": {
        "id": "kuwait",
        "name": "Kuwait",
        "region": "mideast",
        "iso_numeric": "414",
        "iso_alpha3": "KWT",
        "office": "Trademarks and Patent Department, Ministry of Commerce and Industry",
        "filing_language": "Arabic",
        "routes": ["Paris", "PCT National Phase"],
    },
    "oman": {
        "id": "oman",
        "name": "Oman",
        "region": "mideast",
        "iso_numeric": "512",
        "iso_alpha3": "OMN",
        "office": "Ministry of Commerce, Industry and Investment Promotion",
        "filing_language": "Arabic (English accepted with translation)",
        "routes": ["Paris", "PCT National Phase"],
    },
    "qatar": {
        "id": "qatar",
        "name": "Qatar",
        "region": "mideast",
        "iso_numeric": "634",
        "iso_alpha3": "QAT",
        "office": "Industrial Property Department, Ministry of Commerce and Industry",
        "filing_language": "Arabic",
        "routes": ["Paris", "PCT National Phase"],
    },
    "israel": {
        "id": "israel",
        "name": "Israel",
        "region": "mideast",
        "iso_numeric": "376",
        "iso_alpha3": "ISR",
        "office": "Israel Patent Office (ILPO)",
        "filing_language": "English / Hebrew",
        "routes": ["Paris", "PCT National Phase"],
    },

    # Africa
    "algeria": {
        "id": "algeria",
        "name": "Algeria",
        "region": "africa",
        "iso_numeric": "012",
        "iso_alpha3": "DZA",
        "office": "National Institute of Industrial Property (INAPI)",
        "filing_language": "Arabic / French",
        "routes": ["Paris", "PCT National Phase"],
    },
    "djibouti": {
        "id": "djibouti",
        "name": "Djibouti",
        "region": "africa",
        "iso_numeric": "262",
        "iso_alpha3": "DJI",
        "office": "Office Djiboutien de la Propriété Industrielle et Commerciale (ODPIC)",
        "filing_language": "French",
        "routes": ["Paris", "PCT National Phase"],
    },
    "egypt": {
        "id": "egypt",
        "name": "Egypt",
        "region": "africa",
        "iso_numeric": "818",
        "iso_alpha3": "EGY",
        "office": "Egyptian Patent Office",
        "filing_language": "Arabic",
        "routes": ["Paris", "PCT National Phase"],
    },
    "libya": {
        "id": "libya",
        "name": "Libya",
        "region": "africa",
        "iso_numeric": "434",
        "iso_alpha3": "LBY",
        "office": "Industrial Property Office (Trademarks and Patents Office)",
        "filing_language": "Arabic",
        "routes": ["Paris", "PCT National Phase"],
    },
    "morocco": {
        "id": "morocco",
        "name": "Morocco",
        "region": "africa",
        "iso_numeric": "504",
        "iso_alpha3": "MAR",
        "office": "Moroccan Office of Industrial and Commercial Property (OMPIC)",
        "filing_language": "Arabic / French",
        "routes": ["Paris", "PCT National Phase", "EP Validation"],
    },
    "nigeria": {
        "id": "nigeria",
        "name": "Nigeria",
        "region": "africa",
        "iso_numeric": "566",
        "iso_alpha3": "NGA",
        "office": "Trademarks, Patents and Designs Registry",
        "filing_language": "English",
        "routes": ["Paris", "PCT National Phase"],
    },
    "south africa": {
        "id": "south-africa",
        "name": "South Africa",
        "region": "africa",
        "iso_numeric": "710",
        "iso_alpha3": "ZAF",
        "office": "Companies and Intellectual Property Commission (CIPC)",
        "filing_language": "English",
        "routes": ["Paris", "PCT National Phase"],
    },
    "tunisia": {
        "id": "tunisia",
        "name": "Tunisia",
        "region": "africa",
        "iso_numeric": "788",
        "iso_alpha3": "TUN",
        "office": "National Institute for Standardization and Industrial Property (INNORPI)",
        "filing_language": "Arabic / French",
        "routes": ["Paris", "PCT National Phase", "EP Validation"],
    },

    # Turkey is geographically Eurasian but lives on the Middle East tab in
    # this hub (per Rouse MEA team coverage).
    "turkey": {
        "id": "turkey",
        "name": "Turkey",
        "region": "mideast",
        "iso_numeric": "792",
        "iso_alpha3": "TUR",
        "office": "Turkish Patent and Trademark Office (TÜRKPATENT)",
        "filing_language": "Turkish",
        "routes": ["Paris", "PCT National Phase", "EP Validation"],
    },
}

REGIONAL_SYSTEMS: dict[str, dict] = {
    "aripo": {
        "id": "aripo",
        "name": "ARIPO",
        "long_name": "African Regional Intellectual Property Organisation",
        "region": "africa",
        "host_office_country": "Zimbabwe",
        "host_office_city": "Harare",
        "host_coordinates": [31.0335, -17.8252],   # [lon, lat]
        "member_states_count": 22,
        "treaty": "Harare Protocol",
        "filing_language": "English",
        "coverage": "Regional",
    },
    "oapi": {
        "id": "oapi",
        "name": "OAPI",
        "long_name": "Organisation Africaine de la Propriété Intellectuelle",
        "region": "africa",
        "host_office_country": "Cameroon",
        "host_office_city": "Yaoundé",
        "host_coordinates": [11.5021, 3.848],
        "member_states_count": 17,
        "treaty": "Bangui Agreement",
        "filing_language": "French",
        "coverage": "Regional",
    },
    "gcc": {
        "id": "gcc",
        "name": "GCC Patent Office",
        "long_name": "Gulf Cooperation Council Patent Office",
        "region": "mideast",
        "host_office_country": "Saudi Arabia",
        "host_office_city": "Riyadh",
        "host_coordinates": [46.6753, 24.7136],
        "member_states_count": 6,
        "treaty": "GCC Patent Regulation",
        "filing_language": "Arabic",
        "coverage": "Regional",
    },
}

REGIONAL_KEYS = set(REGIONAL_SYSTEMS.keys())


# ----------------------------------------------------------------------------
# Filename parsing
# ----------------------------------------------------------------------------

def filename_to_key(filename: str) -> str:
    """
    Normalise a source filename to a metadata-table key.

    Examples:
        '2026 - UAE  - Patent.xlsx'            -> 'uae'
        '2025 - Saudi Arabia - Patent.xlsx'    -> 'saudi arabia'
        '2025 - Turkey- Patent.xlsx'           -> 'turkey'      (no space)
        '2025 - South Africa - Patent.xlsx'    -> 'south africa'
        '2026 - ARIPO - Patent.xlsx'           -> 'aripo'
    """
    name = filename.lower().replace(".xlsx", "")
    name = re.sub(r"^\d{4}\s*-\s*", "", name)         # year prefix
    name = re.sub(r"\s*-?\s*patent\s*$", "", name)    # "- Patent" suffix
    name = re.sub(r"\s+", " ", name).strip()
    return name


# ----------------------------------------------------------------------------
# Workbook scanning
# ----------------------------------------------------------------------------

# Phrases that reliably appear in the formality "Document" column and almost
# never in the fee table. Used to find the column by content.
DOC_KEYWORDS = (
    "power of attorney",
    "deed of assignment",
    "commercial certificate",
    "specification and drawings",
    "priority application",
    "pct document",
    "pct request",
    "pct application",
    "form p26",  # South Africa adds this
)


def find_document_column(ws) -> tuple[int | None, int | None]:
    """
    Locate the formality table's 'Document' column by counting cells whose
    value contains a known formality-document keyword. Returns (col, start_row)
    where col is 1-based and start_row is the row of the first match in that
    column.
    """
    candidates: dict[int, list[int]] = {}
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if not v:
                continue
            s = str(v).strip().lower()
            if any(kw in s for kw in DOC_KEYWORDS):
                candidates.setdefault(c, []).append(r)
    if not candidates:
        return None, None
    best_col = max(candidates, key=lambda c: len(candidates[c]))
    if len(candidates[best_col]) < 2:
        return None, None
    start_row = min(candidates[best_col])
    return best_col, start_row


def find_info_panel_column(ws) -> tuple[int | None, int | None]:
    """
    Find the column containing the 'Applicant's name…' info panel. Returns
    (col, start_row) at the first matching cell.
    """
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if not v:
                continue
            s = str(v).strip().lower()
            if "applicant" in s and ("name" in s or "nationality" in s):
                return c, r
    return None, None


# ----------------------------------------------------------------------------
# Normalisation
# ----------------------------------------------------------------------------

NORMALISATIONS = (
    (re.compile(r"\bappostile\b", re.IGNORECASE), "Apostille"),
    (re.compile(r"\bappostille\b", re.IGNORECASE), "Apostille"),
    (re.compile(r"\bcertifed\b", re.IGNORECASE), "Certified"),
    (re.compile(r"\brequried\b", re.IGNORECASE), "required"),
    (re.compile(r"\bsimply copy\b", re.IGNORECASE), "Simple copy"),  # Nigeria typo
    (re.compile(r"\becopy\b"), "eCopy"),  # only the lowercase form; preserve "eCopy" exactly
)


def normalise(value) -> str:
    """Trim, collapse whitespace, scrub Excel artefacts, fix common typos."""
    if value is None:
        return ""
    s = str(value).replace("_x000D_", "")
    s = s.replace("\r\n", " ").replace("\n", " ").replace("\r", " ")
    s = re.sub(r"\s+", " ", s).strip()
    if not s:
        return ""
    for pat, repl in NORMALISATIONS:
        s = pat.sub(repl, s)
    return s


def normalise_required(value) -> str:
    """
    Same as normalise(), but additionally canonicalise bare yes/no so the
    front-end's pill renderer can match exact strings ('Yes', 'No', 'Yes, …').
    """
    s = normalise(value)
    if not s:
        return ""
    low = s.lower()
    if low == "yes":
        return "Yes"
    if low == "no":
        return "No"
    # Capitalise leading 'yes,' / 'no,' but leave the rest of the phrase intact.
    if low.startswith("yes,"):
        return "Yes," + s[4:]
    if low.startswith("no,"):
        return "No," + s[3:]
    return s


# ----------------------------------------------------------------------------
# Last-verified date extraction
# ----------------------------------------------------------------------------

MONTHS = {m: i for i, m in enumerate(
    "jan feb mar apr may jun jul aug sep oct nov dec".split(), start=1
)}


def extract_last_verified(ws, fallback_year: int) -> str:
    """
    Try (in order):
      1. Any datetime/date cell in row 1.
      2. A '(Mon YYYY)' substring in any row-1 cell (e.g. OAPI).
      3. Fallback: <fallback_year>-01-01.
    """
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        if isinstance(v, datetime):
            return v.date().isoformat()
        if isinstance(v, date):
            return v.isoformat()
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        if not v:
            continue
        m = re.search(r"\(([a-z]{3})\w*\s+(\d{4})\)", str(v).lower())
        if m:
            mon = MONTHS.get(m.group(1))
            year = int(m.group(2))
            if mon:
                return f"{year}-{mon:02d}-01"
    return f"{fallback_year}-01-01"


# ----------------------------------------------------------------------------
# Per-file extraction
# ----------------------------------------------------------------------------

def extract_file(path: Path, log: list[str]) -> dict | None:
    key = filename_to_key(path.name)
    is_regional = key in REGIONAL_KEYS
    meta = REGIONAL_SYSTEMS.get(key) or COUNTRIES.get(key)
    if not meta:
        log.append(f"[skip] {path.name}: no metadata mapping for key '{key}'")
        return None

    wb = openpyxl.load_workbook(path, data_only=True, read_only=False)
    ws = wb.active

    year_match = re.match(r"^(\d{4})", path.name)
    fallback_year = int(year_match.group(1)) if year_match else 2025

    doc_col, doc_start = find_document_column(ws)
    if doc_col is None:
        log.append(f"[error] {path.name}: could not locate formality table")
        return None

    formality: list[dict] = []
    for r in range(doc_start, ws.max_row + 1):
        doc = normalise(ws.cell(r, doc_col).value)
        if not doc:
            break
        # Skip an in-table header row if present (e.g. "Application and
        # Formality Document Required" appearing above the data).
        if doc.lower().startswith("application and formality"):
            continue
        formality.append({
            "document": doc,
            "required": normalise_required(ws.cell(r, doc_col + 1).value),
            "format": normalise(ws.cell(r, doc_col + 2).value),
            "authentication": normalise(ws.cell(r, doc_col + 3).value),
            "translation": normalise(ws.cell(r, doc_col + 4).value),
            "deadline": normalise(ws.cell(r, doc_col + 5).value),
        })

    if not formality:
        log.append(f"[error] {path.name}: formality table empty after parse")
        return None

    info_col, info_start = find_info_panel_column(ws)
    information: list[str] = []
    if info_col:
        for r in range(info_start, ws.max_row + 1):
            v = normalise(ws.cell(r, info_col).value)
            if not v:
                break
            information.append(v)
    else:
        log.append(f"[warn] {path.name}: no info panel column detected")

    last_verified = extract_last_verified(ws, fallback_year)

    record = dict(meta)
    record["last_verified"] = last_verified
    record["formality_documents"] = formality
    record["information_required"] = information
    if not is_regional:
        record.setdefault("coverage", "National")
    return record


# ----------------------------------------------------------------------------
# Main
# ----------------------------------------------------------------------------

def main() -> int:
    if not PRICE_LISTS_DIR.exists():
        print(f"ERROR: {PRICE_LISTS_DIR} does not exist", file=sys.stderr)
        return 1

    OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)

    log: list[str] = []
    countries: list[dict] = []
    regional: list[dict] = []
    skipped: list[str] = []

    files = sorted(PRICE_LISTS_DIR.glob("*.xlsx"))
    if not files:
        print(f"ERROR: no .xlsx files in {PRICE_LISTS_DIR}", file=sys.stderr)
        return 1

    for path in files:
        rec = extract_file(path, log)
        if rec is None:
            skipped.append(path.name)
            continue
        key = filename_to_key(path.name)
        if key in REGIONAL_KEYS:
            regional.append(rec)
        else:
            countries.append(rec)

    output = {
        "countries": sorted(countries, key=lambda c: c["name"]),
        "regional_systems": sorted(regional, key=lambda c: c["name"]),
        "metadata": {
            "last_generated": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
            "source": "INTA/Price_Lists/*.xlsx",
            "version": "1.0",
        },
    }

    OUTPUT_FILE.write_text(json.dumps(output, indent=2, ensure_ascii=False))
    LOG_FILE.write_text(("\n".join(log) + "\n") if log else "(no warnings)\n")

    total_rows = sum(len(c["formality_documents"]) for c in countries + regional)
    print(f"Wrote {OUTPUT_FILE.relative_to(REPO_ROOT)}")
    print(f"  Files processed:      {len(files)}")
    print(f"  Countries:            {len(countries)}")
    print(f"  Regional systems:     {len(regional)}")
    print(f"  Total formality rows: {total_rows}")
    print(f"  Warnings:             {len([ln for ln in log if ln])}  (see {LOG_FILE.relative_to(REPO_ROOT)})")
    if skipped:
        print(f"  SKIPPED files (no metadata): {skipped}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
